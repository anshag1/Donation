import csv
import io
import uuid
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.exceptions import NotFoundError, ValidationAppError
from app.core.rbac import require_role
from app.deps import CurrentAdmin, DbDep
from app.models.role import ADMIN, SUPER_ADMIN, TREASURER
from app.repositories import donation_repo, event_repo, organization_repo
from app.services.format_utils import format_inr, format_inr_for_pdf
from app.services.pdf.summary_pdf import MONTH_NAMES, SummaryPdfData, render_summary_pdf

router = APIRouter(prefix="/admin/reports", tags=["admin:reports"])

REPORT_ROLES = (SUPER_ADMIN, ADMIN, TREASURER)

EXPORT_HEADERS = [
    "Date",
    "Donor Name",
    "Mobile Number",
    "Event",
    "Amount (INR)",
    "Status",
    "Purpose",
    "Receipt Number",
    "Donation ID",
]

# A single CSV export is capped at this many rows — well beyond any single
# charitable org's realistic donation volume for v1. If usage ever
# approaches this, paginated/streaming export is the right next step rather
# than silently raising the cap further.
MAX_EXPORT_ROWS = 20_000


@router.get("/export.csv")
def export_donations_csv(
    event_id: uuid.UUID | None = None,
    donor_id: uuid.UUID | None = None,
    status: str | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    q: str | None = None,
    db: Session = DbDep,
    current_admin: CurrentAdmin = Depends(require_role(*REPORT_ROLES)),
) -> StreamingResponse:
    rows, total = donation_repo.list_paginated_admin(
        db,
        current_admin.organization_id,
        page=1,
        page_size=MAX_EXPORT_ROWS,
        event_id=event_id,
        donor_id=donor_id,
        status=status,
        date_from=date_from,
        date_to=date_to,
        q=q,
    )

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(EXPORT_HEADERS)
    for donation, receipt_number, event_title in rows:
        writer.writerow(
            [
                donation.created_at.isoformat(),
                donation.donor_snapshot_json.get("full_name", ""),
                donation.donor_snapshot_json.get("mobile_number", ""),
                event_title or "",
                format_inr(donation.amount_in_paise),
                donation.status,
                donation.purpose or "",
                receipt_number or "",
                str(donation.id),
            ]
        )
    if total > MAX_EXPORT_ROWS:
        writer.writerow([])
        writer.writerow(
            [f"NOTE: {total} donations matched these filters; only the first {MAX_EXPORT_ROWS} are included."]
        )

    buffer.seek(0)
    filename = f"donations-export-{datetime.now().strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export.xlsx")
def export_donations_xlsx(
    event_id: uuid.UUID | None = None,
    donor_id: uuid.UUID | None = None,
    status: str | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    q: str | None = None,
    db: Session = DbDep,
    current_admin: CurrentAdmin = Depends(require_role(*REPORT_ROLES)),
) -> StreamingResponse:
    """Same filters and row cap as export.csv — a presentation-layer
    alternative over the same underlying query, not a separate data path."""
    rows, total = donation_repo.list_paginated_admin(
        db,
        current_admin.organization_id,
        page=1,
        page_size=MAX_EXPORT_ROWS,
        event_id=event_id,
        donor_id=donor_id,
        status=status,
        date_from=date_from,
        date_to=date_to,
        q=q,
    )

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None  # always true for a freshly-created Workbook()
    sheet.title = "Donations"
    sheet.append(EXPORT_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for donation, receipt_number, event_title in rows:
        created_at = donation.created_at.replace(tzinfo=None) if donation.created_at.tzinfo else donation.created_at
        sheet.append(
            [
                created_at,
                donation.donor_snapshot_json.get("full_name", ""),
                donation.donor_snapshot_json.get("mobile_number", ""),
                event_title or "",
                format_inr(donation.amount_in_paise),
                donation.status,
                donation.purpose or "",
                receipt_number or "",
                str(donation.id),
            ]
        )
    if total > MAX_EXPORT_ROWS:
        sheet.append([])
        sheet.append([f"NOTE: {total} donations matched these filters; only the first {MAX_EXPORT_ROWS} are included."])

    for col_index, column_cells in enumerate(sheet.columns, start=1):
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
        sheet.column_dimensions[get_column_letter(col_index)].width = min(max_length + 2, 40)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"donations-export-{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/summary.pdf")
def export_summary_pdf(
    period: str = Query(pattern="^(event|month|year)$"),
    event_id: uuid.UUID | None = None,
    year: int | None = Query(default=None, ge=2000, le=2100),
    month: int | None = Query(default=None, ge=1, le=12),
    db: Session = DbDep,
    current_admin: CurrentAdmin = Depends(require_role(*REPORT_ROLES)),
) -> StreamingResponse:
    organization = organization_repo.get_by_id(db, current_admin.organization_id)
    if organization is None:
        raise NotFoundError("Organization not found")

    if period == "event":
        if event_id is None:
            raise ValidationAppError("event_id is required for an event summary")
        event = event_repo.get_by_id(db, current_admin.organization_id, event_id)
        if event is None:
            raise NotFoundError("Event not found")
        total_amount, total_count = donation_repo.aggregate_totals_for_event(
            db, current_admin.organization_id, event_id
        )
        report_title = f"Event Summary — {event.title}"
        breakdown_headers: list[str] = []
        breakdown_rows: list[list[str]] = []

    elif period == "month":
        if year is None or month is None:
            raise ValidationAppError("year and month are required for a monthly summary")
        total_amount, total_count = donation_repo.aggregate_totals_for_month(
            db, current_admin.organization_id, year=year, month=month
        )
        report_title = f"Monthly Summary — {MONTH_NAMES[month]} {year}"
        breakdown_headers = []
        breakdown_rows = []

    else:
        if year is None:
            raise ValidationAppError("year is required for a yearly summary")
        monthly = donation_repo.aggregate_monthly_breakdown_for_year(db, current_admin.organization_id, year=year)
        total_amount = sum(amount for _, amount, _ in monthly)
        total_count = sum(count for _, _, count in monthly)
        report_title = f"Yearly Summary — {year}"
        breakdown_headers = ["Month", "Amount", "Donations"]
        breakdown_rows = [
            [MONTH_NAMES[month_num], format_inr_for_pdf(amount), str(count)] for month_num, amount, count in monthly
        ]

    pdf_bytes = render_summary_pdf(
        SummaryPdfData(
            organization_name=organization.name,
            report_title=report_title,
            generated_at=datetime.now(),
            total_amount_display=format_inr_for_pdf(total_amount),
            total_count=total_count,
            breakdown_headers=breakdown_headers,
            breakdown_rows=breakdown_rows,
        )
    )
    filename = f"summary-{period}-{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        iter([pdf_bytes]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
